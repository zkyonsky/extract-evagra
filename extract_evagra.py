from pathlib import Path  # Standard Python Module
from openpyxl import load_workbook, Workbook  # pip install openpyxl

#Mengambil semua path file excel pada folder yang dituju
SOURCE_DIR = "Excel" # e.g. r"C:\Users\Username\Desktop\Sample Files"
excel_files = list(Path(SOURCE_DIR).glob("*.xlsx"))

#Mengakses semua file excel dengan perulangan
#Mengakses worksheet dan mengambil nilai evagra pada range B17:I27
#Menyimpat nilai evagra pada sebuah list

daftar_pelatihan = []

for excel_file in excel_files:
    try:
        wb = load_workbook(filename=excel_file, read_only=True, data_only=True)

        tabel = wb["penyelenggaraan"]["B17:I27"]
        pelatihan = excel_file.stem.replace(".xlsx", "")
        daftar_evagra = []
        evagra = {}
        evagra["Pelatihan"] = pelatihan
        for row in tabel :
            for i in range(len(row)):
                evagra[row[0].value] = row[7].value
        daftar_evagra.append(evagra)
        daftar_pelatihan.append(daftar_evagra)
    except Exception:
        pass

#Mengakses worksheet pada file template excel
#Menyisipkan nilai evajar sesuai dengan nama kolom (butir pertanyaan)
#Menyimpan file template sebagai file output

wb = load_workbook(filename="Template_Rekap_Evagra.xlsx")
ws = wb["Sheet1"]

rng = ws[f"B1:AA1"]
i=0
for daftar_evagra in daftar_pelatihan:
    
    for cells in rng:
        i += 1
        for cell in cells:
            
            for evagra in daftar_evagra:
                
                if cell.value in evagra:
                  if (cell.offset(row=i, column=0).value is None):
                     cell.offset(row=i, column=0).value = evagra[cell.value]
                  elif (type(cell.offset(row=i, column=0).value) == float and type(evagra[cell.value]) == float):
                     cell.offset(row=1, column=0).value = (evagra[cell.value]+pengajar[cell.value])/2
                  else: cell.offset(row=i, column=0).value = evagra[cell.value]
            
wb.save("Output_Rekap_Evagra.xlsx")
