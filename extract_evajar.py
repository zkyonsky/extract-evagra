from pathlib import Path  # Standard Python Module
from openpyxl import load_workbook, Workbook  # pip install openpyxl

#Mengambil semua path file excel pada folder yang dituju
SOURCE_DIR = "Excel" # e.g. r"C:\Users\Username\Desktop\Sample Files"
excel_files = list(Path(SOURCE_DIR).glob("*.xlsx"))

#Mengakses semua file excel dengan perulangan
#Mengakses worksheet dan mengambil nilai evajar pada range C18:F30
#Menyimpat nilai evajar pada sebuah list

daftar_pelatihan = []

for excel_file in excel_files:
    try:
        wb = load_workbook(filename=excel_file, read_only=True, data_only=True)

        tabel = wb["Pengajar A"]["C18:F23"]
        pelatihan = excel_file.stem.replace(".xlsx", "")
        daftar_pengajar = []
        pengajar = {}
        pengajar["Pelatihan"] = pelatihan
        for row in tabel :
            for i in range(len(row)):
                pengajar[row[0].value] = row[3].value
        daftar_pengajar.append(pengajar)
        daftar_pelatihan.append(daftar_pengajar)
    except Exception:
        pass

#Mengakses worksheet pada file template excel
#Menyisipkan nilai evajar sesuai dengan nama kolom (nama pengajar)
#Menyimpan file tempalte sebagai file output

wb = load_workbook(filename="Template_Rekap_Evajar.xlsx")
ws = wb["Sheet1"]

rng = ws[f"B1:AG1"]
i=0
for daftar_pengajar in daftar_pelatihan:
    
    for cells in rng:
        i += 1
        for cell in cells:
            
            for pengajar in daftar_pengajar:
                
                if cell.value in pengajar:
                  if (cell.offset(row=i, column=0).value is None):
                     cell.offset(row=i, column=0).value = pengajar[cell.value]
                  elif (type(cell.offset(row=i, column=0).value) == float and type(pengajar[cell.value]) == float):
                     cell.offset(row=1, column=0).value = (pengajar[cell.value]+pengajar[cell.value])/2
                  else: cell.offset(row=i, column=0).value = pengajar[cell.value]
            
wb.save("Output_Rekap_Evajar.xlsx")
